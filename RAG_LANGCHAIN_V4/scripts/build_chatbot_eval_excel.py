from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT_DIR = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT_DIR / "outputs" / "chatbot_evaluation"
RESULTS_CSV = OUTPUT_DIR / "chatbot_ragas_results.csv"
SUMMARY_JSON = OUTPUT_DIR / "chatbot_ragas_summary.json"
XLSX_PATH = OUTPUT_DIR / "DanhGiaChatBot_Ragas.xlsx"
PNG_PATH = OUTPUT_DIR / "DanhGiaChatBot_visualization.png"

GROUP_LABELS = {
    "01_loai_san_pham": "Loại sản phẩm",
    "02_bo_loc_thuong_mai": "Bộ lọc/thương mại",
    "03_thanh_phan_cong_dung": "Thành phần/công dụng",
    "04_ten_san_pham": "Tên sản phẩm",
    "05_follow_up": "Follow-up",
    "06_ngoai_pham_vi": "Ngoài phạm vi",
    "07_edge_case": "Edge case",
}


def safe_text(value: Any, limit: int = 5000) -> str:
    if pd.isna(value):
        return ""
    text = str(value)
    return text if len(text) <= limit else f"{text[: limit - 3]}..."


def percent(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def style_header(ws, row: int, max_col: int, fill: str = "244C5A") -> None:
    for cell in ws[row][:max_col]:
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def set_widths(ws, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def build_visualization(summary: dict[str, Any]) -> None:
    group_df = pd.DataFrame(summary["by_group"])
    group_df["label"] = group_df["group"].map(GROUP_LABELS)

    fig, axes = plt.subplots(2, 2, figsize=(16, 10))
    fig.suptitle("Đánh giá Chatbot RAG bằng Ragas - 100 câu hỏi", fontsize=18, fontweight="bold")

    axes[0, 0].barh(group_df["label"], group_df["overall_pass_rate"], color="#2F7D6D")
    axes[0, 0].set_xlim(0, 1)
    axes[0, 0].set_title("Tỉ lệ đạt rule tổng hợp theo nhóm")
    axes[0, 0].xaxis.set_major_formatter(lambda value, _: f"{value:.0%}")

    axes[0, 1].barh(group_df["label"], group_df["avg_latency_sec"], color="#C67A3D")
    axes[0, 1].set_title("Latency trung bình theo nhóm (giây)")

    pass_count = int(round(summary["overall_rule_pass_rate"] * summary["total_questions"]))
    fail_count = summary["total_questions"] - pass_count
    axes[1, 0].pie(
        [pass_count, fail_count],
        labels=["Đạt", "Không đạt"],
        autopct="%1.0f%%",
        colors=["#2F7D6D", "#B84A4A"],
        startangle=90,
    )
    axes[1, 0].set_title("Tổng quan đạt / không đạt")

    axes[1, 1].bar(group_df["label"], group_df["avg_semantic_similarity"], color="#556B8E")
    axes[1, 1].set_ylim(0, max(0.65, group_df["avg_semantic_similarity"].max() + 0.1))
    axes[1, 1].set_title("Ragas semantic similarity trung bình")
    axes[1, 1].tick_params(axis="x", rotation=35)

    for ax in axes.flatten():
        ax.grid(axis="x", alpha=0.25)

    plt.tight_layout(rect=(0, 0, 1, 0.95))
    fig.savefig(PNG_PATH, dpi=180, bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.read_csv(RESULTS_CSV)
    summary = json.loads(SUMMARY_JSON.read_text(encoding="utf-8"))
    group_df = pd.DataFrame(summary["by_group"])
    group_df["label"] = group_df["group"].map(GROUP_LABELS)
    build_visualization(summary)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"] = "Đánh giá Chatbot RAG bằng Ragas - 100 câu hỏi"
    ws["A1"].fill = PatternFill("solid", fgColor="244C5A")
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:J2")
    ws["A2"] = "Nguồn dữ liệu: chatbot_ragas_results.csv | Thời điểm chạy: 2026-06-02"
    ws["A2"].font = Font(color="56616A")

    kpis = [
        ("Tổng số câu hỏi", summary["total_questions"], "0"),
        ("HTTP thành công", summary["successful_http"], "0"),
        ("Độ chính xác routing", summary["route_accuracy"], "0.0%"),
        ("Tỉ lệ đủ sản phẩm", summary["product_count_pass_rate"], "0.0%"),
        ("Tỉ lệ đạt rule tổng hợp", summary["overall_rule_pass_rate"], "0.0%"),
        ("Điểm keyword trung bình", summary["avg_keyword_score"], "0.0%"),
        ("Ragas semantic similarity TB", summary["avg_semantic_similarity"], "0.000"),
        ("Latency trung bình (giây)", summary["avg_latency_sec"], "0.00"),
        ("P95 latency (giây)", summary["p95_latency_sec"], "0.00"),
    ]
    ws.append([])
    ws.append(["Chỉ số", "Giá trị", "Diễn giải"])
    style_header(ws, 4, 3)
    explanations = {
        "Độ chính xác routing": "Tỉ lệ route thực tế khớp route kỳ vọng product_rag/chitchat.",
        "Tỉ lệ đủ sản phẩm": "Tỉ lệ câu product_rag trả đủ số sản phẩm tối thiểu kỳ vọng.",
        "Tỉ lệ đạt rule tổng hợp": "HTTP 200, route đúng, đủ sản phẩm, keyword_score đạt ngưỡng.",
        "Ragas semantic similarity TB": "Điểm Ragas embedding similarity giữa response và reference rút gọn.",
    }
    for label, value, number_format in kpis:
        ws.append([label, value, explanations.get(label, "")])
        ws.cell(ws.max_row, 2).number_format = number_format

    start = 15
    headers = [
        "Nhóm",
        "Số câu",
        "Route acc",
        "Đủ SP",
        "Overall",
        "Keyword",
        "Semantic",
        "Latency TB",
        "SP TB",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(start, col, header)
    style_header(ws, start, len(headers))
    for row_idx, item in enumerate(summary["by_group"], start=start + 1):
        values = [
            GROUP_LABELS.get(item["group"], item["group"]),
            item["questions"],
            item["route_accuracy"],
            item["product_pass_rate"],
            item["overall_pass_rate"],
            item["avg_keyword_score"],
            item["avg_semantic_similarity"],
            item["avg_latency_sec"],
            item["avg_product_count"],
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col, value)
        for col in range(3, 7):
            ws.cell(row_idx, col).number_format = "0.0%"
        ws.cell(row_idx, 7).number_format = "0.000"
        ws.cell(row_idx, 8).number_format = "0.00"
        ws.cell(row_idx, 9).number_format = "0.00"
    add_table(ws, "BangDashboardNhom", f"A{start}:I{start + len(summary['by_group'])}")

    ws["K4"] = "Kết quả"
    ws["L4"] = "Số câu"
    ws["K5"] = "Đạt"
    ws["L5"] = int(df["overall_rule_pass"].sum())
    ws["K6"] = "Không đạt"
    ws["L6"] = int((~df["overall_rule_pass"]).sum())
    style_header(ws, 4, 12, "5A6B2E")

    chart = BarChart()
    chart.title = "Tỉ lệ đạt theo nhóm"
    chart.y_axis.title = "Tỉ lệ"
    chart.x_axis.title = "Nhóm"
    data = Reference(ws, min_col=3, max_col=5, min_row=start, max_row=start + len(summary["by_group"]))
    cats = Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(summary["by_group"]))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 17
    ws.add_chart(chart, "K8")

    latency_chart = BarChart()
    latency_chart.title = "Latency trung bình theo nhóm"
    latency_data = Reference(ws, min_col=8, min_row=start, max_row=start + len(summary["by_group"]))
    latency_chart.add_data(latency_data, titles_from_data=True)
    latency_chart.set_categories(cats)
    latency_chart.height = 9
    latency_chart.width = 17
    ws.add_chart(latency_chart, "K25")

    donut = DoughnutChart()
    donut.title = "Đạt / Không đạt"
    donut.add_data(Reference(ws, min_col=12, min_row=4, max_row=6), titles_from_data=True)
    donut.set_categories(Reference(ws, min_col=11, min_row=5, max_row=6))
    donut.height = 8
    donut.width = 12
    ws.add_chart(donut, "A25")
    ws.add_image(Image(str(PNG_PATH)), "A43")

    ws.freeze_panes = "A4"
    set_widths(ws, [24, 16, 62, 14, 14, 14, 14, 14, 12, 4, 20, 12])

    ws_group = wb.create_sheet("NhomDanhGia")
    ws_group.sheet_view.showGridLines = False
    group_headers = [
        "Group",
        "Tên nhóm",
        "Số câu",
        "Route accuracy",
        "Product pass rate",
        "Overall pass rate",
        "Keyword score TB",
        "Ragas semantic TB",
        "Latency TB",
        "Số sản phẩm TB",
    ]
    ws_group.append(group_headers)
    for item in summary["by_group"]:
        ws_group.append(
            [
                item["group"],
                GROUP_LABELS.get(item["group"], item["group"]),
                item["questions"],
                item["route_accuracy"],
                item["product_pass_rate"],
                item["overall_pass_rate"],
                item["avg_keyword_score"],
                item["avg_semantic_similarity"],
                item["avg_latency_sec"],
                item["avg_product_count"],
            ]
        )
    style_header(ws_group, 1, len(group_headers))
    for row in ws_group.iter_rows(min_row=2, min_col=4, max_col=8):
        for cell in row:
            cell.number_format = "0.0%"
    for row in ws_group.iter_rows(min_row=2, min_col=9, max_col=10):
        for cell in row:
            cell.number_format = "0.00"
    add_table(ws_group, "BangNhomDanhGia", f"A1:J{ws_group.max_row}")
    ws_group.freeze_panes = "A2"
    set_widths(ws_group, [26, 24, 10, 16, 18, 18, 18, 18, 14, 16])

    detail_cols = [
        ("id", "ID", 10),
        ("group", "Nhóm", 22),
        ("query", "Câu hỏi", 44),
        ("setup_turns", "Setup turns", 34),
        ("expected_route", "Route kỳ vọng", 16),
        ("actual_route", "Route thực tế", 16),
        ("route_correct", "Route đúng", 14),
        ("min_products", "Min SP", 10),
        ("product_count", "Số SP", 10),
        ("product_count_pass", "Đủ SP", 12),
        ("source_count", "Số nguồn", 12),
        ("keyword_score", "Keyword score", 16),
        ("overall_rule_pass", "Overall pass", 16),
        ("confidence", "Confidence", 14),
        ("latency_sec", "Latency sec", 14),
        ("status_code", "HTTP", 10),
        ("semantic_similarity", "Ragas semantic", 18),
        ("rewrite_question", "Rewrite question", 42),
        ("product_names", "Tên sản phẩm", 55),
        ("answer", "Câu trả lời", 65),
        ("expected_terms", "Expected terms", 30),
        ("error", "Error", 25),
    ]
    ws_detail = wb.create_sheet("ChiTiet100Cau")
    ws_detail.sheet_view.showGridLines = False
    ws_detail.append([label for _, label, _ in detail_cols])
    for _, row in df.iterrows():
        ws_detail.append([safe_text(row[col], 2500) for col, _, _ in detail_cols])
    style_header(ws_detail, 1, len(detail_cols))
    add_table(ws_detail, "BangChiTiet100Cau", f"A1:V{ws_detail.max_row}")
    ws_detail.freeze_panes = "C2"
    set_widths(ws_detail, [width for _, _, width in detail_cols])
    for row in ws_detail.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in ["L", "M", "N", "Q"]:
        for cell in ws_detail[col][1:]:
            cell.number_format = "0.000"
    for col in ["O"]:
        for cell in ws_detail[col][1:]:
            cell.number_format = "0.00"

    red_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    green_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    for col in ["G", "J", "M"]:
        ws_detail.conditional_formatting.add(
            f"{col}2:{col}{ws_detail.max_row}",
            CellIsRule(operator="equal", formula=["FALSE"], fill=red_fill),
        )
        ws_detail.conditional_formatting.add(
            f"{col}2:{col}{ws_detail.max_row}",
            CellIsRule(operator="equal", formula=["TRUE"], fill=green_fill),
        )

    ws_fail = wb.create_sheet("CasesCanXem")
    ws_fail.sheet_view.showGridLines = False
    fail_df = df[~df["overall_rule_pass"]].copy()
    fail_df["failure_reason"] = fail_df.apply(
        lambda row: "; ".join(
            reason
            for reason in [
                "" if row["status_code"] == 200 else f"HTTP {row['status_code']}",
                "" if row["route_correct"] else f"route {row['actual_route']} thay vì {row['expected_route']}",
                ""
                if row["product_count_pass"]
                else f"chỉ có {row['product_count']}/{row['min_products']} sản phẩm kỳ vọng",
                "" if row["keyword_score"] >= 0.5 else f"keyword_score={row['keyword_score']}",
            ]
            if reason
        )
        or "Cần kiểm tra thủ công",
        axis=1,
    )
    fail_cols = [
        "id",
        "group",
        "query",
        "failure_reason",
        "expected_route",
        "actual_route",
        "product_count",
        "keyword_score",
        "semantic_similarity",
        "answer",
        "product_names",
    ]
    ws_fail.append(
        [
            "ID",
            "Nhóm",
            "Câu hỏi",
            "Lý do fail",
            "Route kỳ vọng",
            "Route thực tế",
            "Số SP",
            "Keyword score",
            "Ragas semantic",
            "Câu trả lời",
            "Tên sản phẩm",
        ]
    )
    for _, row in fail_df.iterrows():
        ws_fail.append([safe_text(row[col], 2500) for col in fail_cols])
    style_header(ws_fail, 1, len(fail_cols), "7A3E3E")
    add_table(ws_fail, "BangCasesCanXem", f"A1:K{ws_fail.max_row}")
    ws_fail.freeze_panes = "A2"
    set_widths(ws_fail, [10, 22, 44, 50, 16, 16, 10, 16, 18, 65, 55])
    for row in ws_fail.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws_ragas = wb.create_sheet("RagasInput")
    ws_ragas.sheet_view.showGridLines = False
    ws_ragas.append(["ID", "User input", "Response", "Reference", "Retrieved contexts"])
    for _, row in df.iterrows():
        ws_ragas.append(
            [
                row["id"],
                safe_text(row["query"], 800),
                safe_text(row["answer"], 5000),
                safe_text(row["reference"], 1500),
                safe_text(row["retrieved_contexts"], 5000),
            ]
        )
    style_header(ws_ragas, 1, 5)
    add_table(ws_ragas, "BangRagasInput", f"A1:E{ws_ragas.max_row}")
    ws_ragas.freeze_panes = "A2"
    set_widths(ws_ragas, [10, 45, 70, 60, 70])
    for row in ws_ragas.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    thin = Side(style="thin", color="D9DEE3")
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin)

    wb.save(XLSX_PATH)
    verified = load_workbook(XLSX_PATH, read_only=False)
    assert set(verified.sheetnames) >= {
        "Dashboard",
        "NhomDanhGia",
        "ChiTiet100Cau",
        "CasesCanXem",
        "RagasInput",
    }
    print(json.dumps({"xlsx": str(XLSX_PATH), "png": str(PNG_PATH), "sheets": verified.sheetnames}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
